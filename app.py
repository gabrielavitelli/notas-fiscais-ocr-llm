"""
Notas Fiscais — Dashboard para extração estruturada de notas fiscais (OCR + IA).
Uma única aba: upload → processamento → resultados (filtros e exportação CSV).
Chaves de API no arquivo .env.
"""
import os

# Headless para OpenCV/DocTR na nuvem (antes de qualquer import que use cv2)
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
os.environ.setdefault("OPENCV_OPENCL_RUNTIME", "")
os.environ.setdefault("DISPLAY", "")

from pathlib import Path

_env_path = Path(__file__).resolve().parent / ".env"
if _env_path.exists():
    try:
        from dotenv import load_dotenv
        load_dotenv(_env_path)
    except (ImportError, UnicodeDecodeError):
        pass
    except Exception:
        pass  # .env com encoding errado (ex. UTF-16): na nuvem use Secrets; no PC salve .env em UTF-8

import json
import io
import re
import tempfile
import time
import csv as csv_module
import unicodedata
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# nf_ocr é importado só ao processar (evita travar o deploy com deps pesadas)


def _load_secrets_to_env():
    """Coloca GROQ_API_KEY e HF_TOKEN de st.secrets em os.environ (Streamlit Cloud)."""
    try:
        sk = getattr(st, "secrets", None)
        if sk is None:
            return
        sources = [sk]
        try:
            if hasattr(sk, "get") and sk.get("default") is not None:
                sources.append(sk["default"])
        except (KeyError, TypeError):
            pass
        for name, env_key in [("GROQ_API_KEY", "GROQ_API_KEY"), ("groq_api_key", "GROQ_API_KEY"), ("HF_TOKEN", "HF_TOKEN"), ("hf_token", "HF_TOKEN")]:
            for src in sources:
                try:
                    v = src[name] if hasattr(src, "__getitem__") else (src.get(name) if hasattr(src, "get") else None)
                    if isinstance(v, str) and v.strip():
                        os.environ[env_key] = v.strip()
                        break
                except (KeyError, TypeError, AttributeError):
                    pass
    except Exception:
        pass


def _resolve_gcp_project_id_for_secrets():
    """Projeto onde o Secret Manager procura os segredos (deve ser o mesmo do secret)."""
    for k in ("GOOGLE_CLOUD_PROJECT", "GCP_PROJECT", "GCLOUD_PROJECT"):
        v = (os.environ.get(k) or "").strip()
        if v:
            return v
    # Service account JSON traz o project_id — evita erro quando o secret não está em "climaticsystem"
    cred_path = (os.environ.get("GOOGLE_APPLICATION_CREDENTIALS") or "").strip().strip('"')
    if cred_path:
        try:
            p = Path(cred_path)
            if p.is_file():
                data = json.loads(p.read_text(encoding="utf-8"))
                pid = data.get("project_id")
                if isinstance(pid, str) and pid.strip():
                    return pid.strip()
        except Exception:
            pass
    return "climaticsystem"


def _secret_manager_service_client():
    """Cria cliente do Secret Manager (import compatível com vários ambientes)."""
    # 1) Caminho “oficial” do pacote google-cloud-secret-manager
    try:
        from google.cloud import secretmanager

        return secretmanager.SecretManagerServiceClient()
    except ImportError:
        pass
    # 2) Import direto (às vezes resolve conflito de namespace google.cloud)
    try:
        from google.cloud.secretmanager_v1 import SecretManagerServiceClient

        return SecretManagerServiceClient()
    except ImportError as e:
        raise ImportError(
            "Instale o cliente do Secret Manager no MESMO Python que roda o Streamlit: "
            "python -m pip install google-cloud-secret-manager"
        ) from e


def _get_secret_from_gcp(secret_id, project=None, version="latest"):
    """Lê segredo no Google Secret Manager. Retorna string ou None."""
    try:
        project_id = project or _resolve_gcp_project_id_for_secrets()
        client = _secret_manager_service_client()
        name = f"projects/{project_id}/secrets/{secret_id}/versions/{version}"
        response = client.access_secret_version(request={"name": name})
        value = response.payload.data.decode("UTF-8").strip()
        return value or None
    except Exception as e:
        try:
            os.environ["GCP_SECRET_MANAGER_LAST_ERROR"] = f"{type(e).__name__}: {str(e)[:280]}"
        except Exception:
            pass
        return None


def _load_gcp_secret_manager_to_env():
    """Fallback: carrega GROQ/HF do Google Secret Manager quando env/secrets não vierem preenchidos."""
    try:
        os.environ.pop("GCP_SECRET_MANAGER_LAST_ERROR", None)
        project_id = _resolve_gcp_project_id_for_secrets()
        try:
            os.environ["GCP_SECRET_MANAGER_PROJECT_USED"] = project_id
        except Exception:
            pass
        # Ordem tentativa por token para facilitar reaproveitar segredos já existentes no projeto.
        candidates = {
            "GROQ_API_KEY": ["GROQ_API_KEY", "groq_api_key", "GROQ-API-KEY", "groq-api-key"],
            "HF_TOKEN": ["HF_TOKEN", "hf_token", "HUGGINGFACE_TOKEN", "huggingface_token"],
        }
        for env_key, secret_ids in candidates.items():
            if os.environ.get(env_key, "").strip():
                continue
            for sid in secret_ids:
                value = _get_secret_from_gcp(sid, project=project_id)
                if isinstance(value, str) and value.strip():
                    os.environ[env_key] = value.strip()
                    break
    except Exception:
        pass


VERSION = "1.0.0"
LAST_UPDATE = "2025-03"
APP_LOGIN_USER = "geoinfra"
APP_LOGIN_PASSWORD = "geoinfra"

# Diretório padrão: CSV, estado e log num só lugar (na nuvem, se read-only, usa temp)
try:
    _base = Path(__file__).resolve().parent
except Exception:
    _base = Path(".")
NF_DADOS_DIR = _base / "nf_dados"
_fallback_dados_dir = None


def _get_dados_dir():
    """Retorna o diretório de dados (nf_dados ou temp se não for possível escrever na pasta do app)."""
    global _fallback_dados_dir
    if _fallback_dados_dir is not None:
        return _fallback_dados_dir
    try:
        NF_DADOS_DIR.mkdir(parents=True, exist_ok=True)
        # Testa se consegue escrever (na nuvem a pasta do app pode ser read-only)
        (NF_DADOS_DIR / ".write_test").write_text("ok", encoding="utf-8")
        (NF_DADOS_DIR / ".write_test").unlink(missing_ok=True)
        return NF_DADOS_DIR
    except Exception:
        try:
            _fallback_dados_dir = Path(tempfile.gettempdir()) / "notas_fiscais_nf_dados"
            _fallback_dados_dir.mkdir(parents=True, exist_ok=True)
            return _fallback_dados_dir
        except Exception:
            return NF_DADOS_DIR  # último recurso, pode falhar ao salvar


def _ensure_dados_dir():
    """Garante que o diretório de dados existe (chama _get_dados_dir())."""
    _get_dados_dir()


def _state_file():
    return _get_dados_dir() / "notas_fiscais_state.json"


def _log_duvidas_file():
    return _get_dados_dir() / "notas_fiscais_duvidas_erros.json"


def _default_csv_path():
    return _get_dados_dir() / "nf_extraidas.csv"

STYLE = """
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
  html, body, [data-testid="stAppViewContainer"] { font-family: 'Inter', sans-serif; background: #F8FAFC; }
  /* Sidebar: fundo claro #F1F5F9 → texto bem escuro para contraste */
  section[data-testid="stSidebar"] { background: #F1F5F9 !important; }
  section[data-testid="stSidebar"] > div:first-child {
    width: 360px !important;
    min-width: 360px !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
  }
  section[data-testid="stSidebar"] .stMarkdown,
  section[data-testid="stSidebar"] .stMarkdown p,
  section[data-testid="stSidebar"] .stMarkdown h1,
  section[data-testid="stSidebar"] .stMarkdown h2,
  section[data-testid="stSidebar"] .stMarkdown h3,
  section[data-testid="stSidebar"] .stMarkdown span,
  section[data-testid="stSidebar"] .stCaption,
  section[data-testid="stSidebar"] .stCaption span,
  section[data-testid="stSidebar"] label,
  section[data-testid="stSidebar"] span,
  section[data-testid="stSidebar"] p,
  section[data-testid="stSidebar"] div,
  section[data-testid="stSidebar"] [data-testid="stRadio"] label,
  section[data-testid="stSidebar"] [data-testid="stRadio"] span,
  section[data-testid="stSidebar"] [data-testid="stRadio"] p,
  section[data-testid="stSidebar"] [data-testid="stRadio"] div,
  section[data-testid="stSidebar"] * { color: #000000 !important; }
  section[data-testid="stSidebar"] a { color: #000000 !important; }
  /* Área principal (Início, Upload, Recente): texto preto em fundo claro */
  [data-testid="stAppViewContainer"] .stMarkdown,
  [data-testid="stAppViewContainer"] .stMarkdown p,
  [data-testid="stAppViewContainer"] .stMarkdown li,
  [data-testid="stAppViewContainer"] .stMarkdown h1,
  [data-testid="stAppViewContainer"] .stMarkdown h2,
  [data-testid="stAppViewContainer"] .stMarkdown h3,
  [data-testid="stAppViewContainer"] .stMarkdown h4,
  [data-testid="stAppViewContainer"] .stMarkdown h5,
  [data-testid="stAppViewContainer"] .stMarkdown h6,
  [data-testid="stAppViewContainer"] .stMarkdown span,
  [data-testid="stAppViewContainer"] .stCaption,
  [data-testid="stAppViewContainer"] .stCaption span,
  [data-testid="stAppViewContainer"] .stCaption div,
  [data-testid="stVerticalBlock"] .stMarkdown,
  [data-testid="stVerticalBlock"] .stMarkdown p,
  [data-testid="stVerticalBlock"] .stCaption,
  [data-testid="stVerticalBlock"] .stCaption span { color: #000000 !important; }
  section.main .stMarkdown, section.main .stMarkdown p, section.main .stMarkdown li, section.main .stMarkdown h1,
  section.main .stMarkdown h2, section.main .stMarkdown h3, section.main .stMarkdown h4,
  section.main .stMarkdown h5, section.main .stMarkdown h6,
  section.main .stCaption, section.main .stCaption span { color: #000000 !important; }
  .nf-card { background: #FFFFFF; border-radius: 16px; padding: 1.5rem; box-shadow: 0 1px 3px rgba(0,0,0,0.06); border: 1px solid #E2E8F0; margin-bottom: 1rem; color: #000000 !important; }
  .nf-card strong { color: #000000 !important; }
  .nf-dropzone { border: 2px dashed #E2E8F0; border-radius: 16px; padding: 2.5rem; text-align: center; background: #F8FAFC; color: #000000 !important; }
  .nf-badge { display: inline-block; padding: 0.25rem 0.75rem; border-radius: 9999px; font-size: 0.75rem; font-weight: 600; color: #000000 !important; }
  .nf-badge-finalizado { background: #D1FAE5; color: #065f46 !important; }
  .nf-badge-erro { background: #FEE2E2; color: #991b1b !important; }
  .nf-badge-enviado { background: #E2E8F0; color: #334155 !important; }
  .nf-badge-ja_incluido { background: #FEF3C7; color: #92400E !important; }
  .nf-metric { background: #FFFFFF; border-radius: 12px; padding: 1rem; border: 1px solid #E2E8F0; text-align: center; color: #000000 !important; }
  .nf-metric-value { font-size: 1.5rem; font-weight: 700; color: #2563EB !important; }
  .nf-metric-label { font-size: 0.75rem; color: #000000 !important; margin-top: 0.25rem; }
  .nf-app-header { display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 1rem; margin-bottom: 1rem; }
  .nf-app-title { font-size: 1.5rem; font-weight: 700; color: #000000 !important; margin: 0; }
  .nf-app-desc { font-size: 0.875rem; color: #000000 !important; margin: 0.25rem 0 0 0; }
  .nf-status { display: inline-flex; align-items: center; padding: 0.35rem 0.75rem; border-radius: 9999px; font-size: 0.75rem; font-weight: 500; background: #D1FAE5; color: #059669 !important; }
  .nf-postit { background: #FEF3C7; border-radius: 10px; padding: 0.75rem 1rem; border-left: 4px solid #F59E0B; font-size: 0.875rem; margin-bottom: 0.5rem; color: #000000 !important; }
  .stButton > button[kind="primary"] { background: #2563EB !important; color: #FFFFFF !important; border: none !important; font-weight: 600 !important; border-radius: 10px !important; }
  .stFormSubmitButton > button { background: #2563EB !important; color: #FFFFFF !important; border: none !important; font-weight: 600 !important; border-radius: 10px !important; }
  .stFormSubmitButton > button span { color: #FFFFFF !important; }
  /* Widgets na área principal: labels em preto */
  section.main [data-testid="stRadio"] label, section.main [data-testid="stRadio"] span,
  section.main [data-testid="stSelectbox"] label, section.main [data-testid="stSelectbox"] span,
  section.main .stCheckbox label, section.main .stDateInput label,
  [data-testid="stAppViewContainer"] [data-testid="stRadio"] label,
  [data-testid="stAppViewContainer"] [data-testid="stRadio"] span,
  [data-testid="stAppViewContainer"] [data-testid="stSelectbox"] label,
  [data-testid="stAppViewContainer"] [data-testid="stSelectbox"] span,
  [data-testid="stAppViewContainer"] .stCheckbox label,
  [data-testid="stAppViewContainer"] .stDateInput label { color: #000000 !important; }
  section.main [data-testid="stProgress"] span, section.main [data-testid="stProgress"] label { color: #000000 !important; }
  section.main [data-testid="stProgress"] * { color: #000000 !important; }
  /* Nome do arquivo (lista de enviados) em preto */
  [data-testid="stAppViewContainer"] [data-testid="stFileUploader"] label,
  [data-testid="stAppViewContainer"] [data-testid="stFileUploader"] span,
  [data-testid="stAppViewContainer"] [data-testid="stFileUploader"] div,
  [data-testid="stAppViewContainer"] [data-testid="stFileUploader"] p,
  [data-testid="stAppViewContainer"] [data-testid="stFileUploader"] * { color: #000000 !important; }
  /* Texto da área de drop (Drag and drop / Limit 200MB) em BRANCO — depois do preto para ganhar especificidade */
  [data-testid="stAppViewContainer"] [data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"],
  [data-testid="stAppViewContainer"] [data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"] *,
  [data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"],
  [data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"] *,
  /* Fallback: 1º bloco do uploader costuma ser o dropzone (Streamlit pode mudar testid) */
  [data-testid="stFileUploader"] > div:first-child,
  [data-testid="stFileUploader"] > div:first-child *,
  [data-testid="stAppViewContainer"] [data-testid="stFileUploader"] > div:first-child,
  [data-testid="stAppViewContainer"] [data-testid="stFileUploader"] > div:first-child * { color: #FFFFFF !important; fill: #FFFFFF !important; }
  [data-testid="stAppViewContainer"] [data-testid="stProgress"] span,
  [data-testid="stAppViewContainer"] [data-testid="stProgress"] label,
  [data-testid="stAppViewContainer"] [data-testid="stProgress"] * { color: #000000 !important; }
  /* Spinner e texto "Processando… aguarde." / "Iniciando…" em preto */
  [data-testid="stSpinner"] *,
  [data-testid="stSpinner"] { color: #000000 !important; }
  [data-testid="stAppViewContainer"] [data-testid="stSpinner"],
  [data-testid="stAppViewContainer"] [data-testid="stSpinner"] * { color: #000000 !important; }
  /* Layout estilo dashboard: cards e coluna direita */
  .nf-card-panel { background: #FFFFFF; border-radius: 12px; padding: 1.25rem; box-shadow: 0 1px 3px rgba(0,0,0,0.08); border: 1px solid #E2E8F0; margin-bottom: 1rem; }
  .nf-card-panel h4 { margin: 0 0 0.75rem 0; font-size: 1rem; color: #000000; }
  .nf-upload-hero { border: 2px dashed #CBD5E1; border-radius: 16px; padding: 2.5rem; text-align: center; background: #F8FAFC; color: #475569; margin: 0.75rem 0; }
  .nf-upload-hero .cloud { font-size: 2.5rem; margin-bottom: 0.5rem; }
  .nf-sidebar-footer {
    margin-top: 2rem;
    padding: 0.85rem;
    border: 1px solid #E2E8F0;
    border-radius: 10px;
    background: #FFFFFF;
    font-size: 0.82rem;
    color: #0F172A;
    line-height: 1.45;
  }
  .nf-header-right { display: flex; align-items: center; gap: 0.75rem; flex-wrap: wrap; }
  /* Menu lateral: mais espaço e destaque no item ativo */
  section[data-testid="stSidebar"] hr {
    margin: 0.7rem 0 1rem 0 !important;
    border-color: #CBD5E1 !important;
  }
  .nf-nav-title {
    margin: 0.15rem 0 0.35rem 0;
    font-size: 1.45rem;
    font-weight: 800;
    letter-spacing: 0.2px;
    color: #0F172A !important;
  }
  section[data-testid="stSidebar"] [data-testid="stRadio"] { gap: 0.75rem !important; }
  section[data-testid="stSidebar"] [data-testid="stRadio"] > label {
    margin-bottom: 0.5rem !important;
    padding: 1rem 1rem !important;
    border-radius: 14px !important;
    border: 1px solid #CBD5E1 !important;
    background: #FFFFFF !important;
    transition: all .18s ease !important;
    font-weight: 700 !important;
    font-size: 1.02rem !important;
    line-height: 1.35 !important;
  }
  section[data-testid="stSidebar"] [data-testid="stRadio"] > label:hover {
    background: rgba(37, 99, 235, 0.08) !important;
    border-color: #93C5FD !important;
  }
  .nf-nav-helper {
    margin: 0.25rem 0 1rem 0;
    color: #475569 !important;
    font-size: 0.92rem;
  }
  .nf-result-card {
    border: 1px solid #E2E8F0;
    border-radius: 12px;
    padding: 0.8rem 0.9rem;
    margin-bottom: 0.6rem;
    background: #0F172A;
  }
  .nf-result-card-title {
    font-size: 0.88rem;
    font-weight: 600;
    color: #FFFFFF;
    margin-bottom: 0.25rem;
  }
  .nf-result-card-sub {
    font-size: 0.78rem;
    color: #E2E8F0;
    margin-bottom: 0.5rem;
  }
  /* Resultados: texto preto em fundo branco */
  [data-testid="stAppViewContainer"] div[data-testid="stVerticalBlock"] > div .stMarkdown,
  [data-testid="stAppViewContainer"] div[data-testid="stVerticalBlock"] .stCaption,
  [data-testid="stAppViewContainer"] div[data-testid="stVerticalBlock"] label,
  [data-testid="stAppViewContainer"] div[data-testid="stVerticalBlock"] span,
  [data-testid="stAppViewContainer"] div[data-testid="stVerticalBlock"] p,
  [data-testid="stAppViewContainer"] .stRadio label,
  [data-testid="stAppViewContainer"] .stSelectbox label,
  [data-testid="stAppViewContainer"] .stCheckbox label,
  [data-testid="stAppViewContainer"] .stDateInput label { color: #0F172A !important; }
  /* Tabela de resultados: contorno escuro elegante, sem fundo (texto usa tema) */
  [data-testid="stDataFrame"] {
    border: 1px solid #1e293b !important;
    border-radius: 8px !important;
    box-shadow: 0 1px 3px rgba(30, 41, 59, 0.08) !important;
    overflow: hidden !important;
  }
  [data-testid="stDataFrame"] table {
    border: none !important;
  }
  /* Botão Exportar CSV: texto sempre visível (preto em botão claro) */
  [data-testid="stDownloadButton"] button,
  [data-testid="stDownloadButton"] button span,
  [data-testid="stDownloadButton"] a,
  .stDownloadButton button,
  .stDownloadButton button span,
  [data-testid="stAppViewContainer"] [data-testid="stDownloadButton"],
  [data-testid="stAppViewContainer"] [data-testid="stDownloadButton"] * {
    color: #0F172A !important;
  }
  [data-testid="stDownloadButton"] button { background-color: #E2E8F0 !important; }
</style>
"""


def inject_css():
    st.markdown(STYLE, unsafe_allow_html=True)


def _load_state():
    """Carrega estado persistido (lista_pesquisadores, metrics, results, csv_drive_path, dúvidas/erros)."""
    out = {
        "lista_pesquisadores": [],
        "metrics": {"files_total": 0, "avg_ocr_sec": 0, "success_rate": 100, "fields_detected": 12, "revisar": 0, "verificar": 0, "erros_total": 0},
        "results": [],
        "csv_drive_path": "",
        "log_duvidas_erros": [],
        "ultima_execucao": "",
    }
    try:
        sf = _state_file()
        if sf.exists():
            with open(sf, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                out["lista_pesquisadores"] = data.get("lista_pesquisadores") if isinstance(data.get("lista_pesquisadores"), list) else []
                m = data.get("metrics")
                if isinstance(m, dict):
                    out["metrics"] = {**out["metrics"], **m}
                out["results"] = data.get("results") if isinstance(data.get("results"), list) else []
                out["csv_drive_path"] = str(data.get("csv_drive_path") or "").strip()[:500]
                out["ultima_execucao"] = str(data.get("ultima_execucao") or "").strip()[:50]
        lf = _log_duvidas_file()
        if lf.exists():
            with open(lf, "r", encoding="utf-8") as f:
                log = json.load(f)
            if isinstance(log, list):
                out["log_duvidas_erros"] = log
    except Exception:
        pass
    return out


def _save_state():
    """Persiste estado para não perder ao fechar o app."""
    try:
        _ensure_dados_dir()
        data = {
            "lista_pesquisadores": st.session_state.get("lista_pesquisadores", []),
            "metrics": st.session_state.get("metrics", {}),
            "results": st.session_state.get("results", []),
            "csv_drive_path": (st.session_state.get("csv_drive_path") or "").strip(),
            "ultima_execucao": st.session_state.get("ultima_execucao", ""),
        }
        with open(_state_file(), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=0)
    except Exception:
        pass


def _append_log_duvida(tipo, arquivo, mensagem, detalhe=None):
    """Registra dúvida ou erro para análise em longo prazo."""
    try:
        _ensure_dados_dir()
        log = []
        lf = _log_duvidas_file()
        if lf.exists():
            with open(lf, "r", encoding="utf-8") as f:
                log = json.load(f)
        log.append({
            "data_hora": datetime.now().isoformat(),
            "tipo": tipo,
            "arquivo": arquivo,
            "mensagem": mensagem,
            "detalhe": detalhe,
        })
        with open(lf, "w", encoding="utf-8") as f:
            json.dump(log[-500:], f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def init_session_state():
    if "results" not in st.session_state:
        try:
            loaded = _load_state()
        except Exception:
            loaded = {}
        st.session_state.results = loaded.get("results", [])
        st.session_state.metrics = loaded.get("metrics", {"files_total": 0, "avg_ocr_sec": 0, "success_rate": 100, "fields_detected": 12, "revisar": 0, "verificar": 0, "erros_total": 0})
        st.session_state.lista_pesquisadores = loaded.get("lista_pesquisadores", [])
        st.session_state.csv_drive_path = loaded.get("csv_drive_path", "") or ""
        st.session_state.log_duvidas_erros = loaded.get("log_duvidas_erros", [])
        st.session_state.ultima_execucao = loaded.get("ultima_execucao", "") or ""
    if "processing" not in st.session_state:
        st.session_state.processing = []
    if "metrics" not in st.session_state:
        st.session_state.metrics = {
            "files_total": 0,
            "avg_ocr_sec": 0,
            "success_rate": 100,
            "fields_detected": 12,
            "revisar": 0,
            "verificar": 0,
            "erros_total": 0,
        }
    if "lista_pesquisadores" not in st.session_state:
        st.session_state.lista_pesquisadores = []
    if "csv_drive_path" not in st.session_state:
        st.session_state.csv_drive_path = ""
    if "log_duvidas_erros" not in st.session_state:
        st.session_state.log_duvidas_erros = []
    if "ultima_execucao" not in st.session_state:
        st.session_state.ultima_execucao = ""
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False


@st.cache_resource
def carregar_modelo_doctr():
    # Cache do Hugging Face/Doctr em diretório gravável (na nuvem /tmp)
    try:
        _cache = tempfile.gettempdir()
        _hub = os.path.join(_cache, "hf_hub")
        os.makedirs(_hub, exist_ok=True)
        os.environ["HF_HOME"] = _cache
        os.environ["HUGGINGFACE_HUB_CACHE"] = _hub
    except Exception:
        pass
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    os.environ.setdefault("OPENCV_OPENCL_RUNTIME", "")
    import doctr  # noqa: F401
    from doctr.models import ocr_predictor
    return ocr_predictor(pretrained=True)


def render_header():
    ultima = st.session_state.get("ultima_execucao") or "—"
    st.markdown(f"""
    <div class="nf-app-header">
      <div>
        <h1 class="nf-app-title">Sistema para Cadastro de Notas Fiscais</h1>
        <p class="nf-app-desc">Extração estruturada de notas fiscais usando OCR + IA →</p>
      </div>
      <div class="nf-header-right">
        <span class="nf-status">● Online</span>
        <span style="font-size:0.75rem;color:#0F172A;">v{VERSION}</span>
        <span style="font-size:0.75rem;color:#64748B;">Última execução: {ultima}</span>
      </div>
    </div>
    """, unsafe_allow_html=True)


def require_login():
    if st.session_state.get("authenticated", False):
        return True
    st.markdown("## 🔐 Login")
    st.caption("Acesso restrito ao sistema de notas fiscais.")
    with st.form("login_form_nf"):
        user = st.text_input("Usuário")
        password = st.text_input("Senha", type="password")
        submitted = st.form_submit_button("Entrar", use_container_width=True)
    if submitted:
        if user == APP_LOGIN_USER and password == APP_LOGIN_PASSWORD:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Usuário ou senha inválidos.")
    return False


def render_sidebar():
    with st.sidebar:
        if st.button("🚪 Sair", use_container_width=True, key="btn_logout"):
            st.session_state.authenticated = False
            st.rerun()
        st.markdown("---")
        st.markdown('<div class="nf-nav-title">Navegação</div>', unsafe_allow_html=True)
        st.markdown('<div class="nf-nav-helper">Escolha uma etapa do fluxo abaixo.</div>', unsafe_allow_html=True)
        st.markdown("---")
        pages = ["☁️ Upload", "📂 Processados", "⚠️ Revisar", "⚙️ Configurações", "ℹ️ Sobre"]
        page_to_index = {"Início": 0, "Processados": 1, "Revisar": 2, "Configurações": 3, "Sobre": 4}
        current = st.session_state.get("page", "Início")
        try:
            idx = min(page_to_index.get(current, 0), len(pages) - 1)
            page = st.radio("Menu", options=pages, index=idx, key="sidebar_nav", label_visibility="collapsed")
        except Exception:
            page = st.radio("Menu", options=pages, key="sidebar_nav", label_visibility="collapsed")
        if "Upload" in page:
            st.session_state.page = "Início"
        elif "Processados" in page:
            st.session_state.page = "Processados"
        elif "Revisar" in page:
            st.session_state.page = "Revisar"
        elif "Configurações" in page:
            st.session_state.page = "Configurações"
        else:
            st.session_state.page = "Sobre"
        st.markdown("---")
        st.markdown('<div class="nf-sidebar-footer"><strong>Extração estruturada</strong><br/>com DocTR e IA.</div>', unsafe_allow_html=True)


def run_processing(progress_placeholder):
    """Processa um único arquivo por chamada. Retorna True se ainda há mais para processar (a página faz rerun e a barra avança)."""
    processing = st.session_state.get("processing", [])
    to_run = [i for i, p in enumerate(processing) if p.get("bytes") and p.get("status") not in ("Finalizado", "Erro")]
    n_total = len(processing)
    if not to_run:
        return False
    i = to_run[0]
    item = processing[i]
    name = item["name"]
    done_before = n_total - len(to_run)

    try:
        model = carregar_modelo_doctr()
    except Exception as e:
        _ensure_dados_dir()
        import traceback
        err_msg = str(e).strip() or repr(e)
        tb_str = traceback.format_exc()
        # Guarda na sessão para não sumir após o rerun
        st.session_state["doctr_last_error"] = {"msg": err_msg, "tb": tb_str}
        st.warning(
            "**DocTR não carregou** — o OCR não está disponível neste ambiente. "
            "CSV e estado ficam em **nf_dados/**. No PC: `pip install -r requirements-local.txt` e `streamlit run app.py`."
        )
        st.error(f"**Erro técnico:** {err_msg}")
        with st.expander("Traceback completo (copie e envie se for reportar)"):
            st.code(tb_str, language="text")
        # Marca como erro e remove bytes para não ficar em loop "Processando..."
        st.session_state.processing[i]["status"] = "Erro: DocTR não carregou"
        st.session_state.processing[i].pop("bytes", None)
        metrics = st.session_state.get("metrics", {})
        metrics["erros_total"] = metrics.get("erros_total", 0) + 1
        st.session_state.metrics = metrics
        _save_state()
        return False
    # Garante que Secrets estão em os.environ antes de chamar nf_ocr (mesmo após rerun)
    _load_secrets_to_env()
    _load_gcp_secret_manager_to_env()
    api_key = os.environ.get("GROQ_API_KEY") or os.environ.get("groq_api_key")
    nomes_pesquisadores = list(st.session_state.get("lista_pesquisadores", []))
    results = list(st.session_state.get("results", []))
    metrics = st.session_state.get("metrics", {})
    csv_drive_path = (st.session_state.get("csv_drive_path") or "").strip()
    use_drive_csv = bool(csv_drive_path)
    ja_incluidos = []

    st.session_state.processing[i]["status"] = "OCR"
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        if use_drive_csv:
            csv_path = Path(csv_drive_path)
            try:
                csv_path.parent.mkdir(parents=True, exist_ok=True)
            except Exception:
                pass
        else:
            _ensure_dados_dir()
            csv_path = _default_csv_path()
        path = tmpdir / (name or f"file_{i}")
        path.write_bytes(item["bytes"])
        t0 = time.time()
        try:
            import nf_ocr
            ok, dados, ocr_text = nf_ocr.processar_arquivo(
                path, model, api_key, sheet_id=None, creds_path=None,
                csv_path=str(csv_path), dry_run=False, nomes_pesquisadores=nomes_pesquisadores, return_text=True,
            )
            elapsed = time.time() - t0
            if ok and dados is not None:
                st.session_state.processing[i]["status"] = "Finalizado"
                rec = nf_ocr.registro_from_dados(dados)
                rec["_elapsed_sec"] = round(elapsed, 1)
                rec["_ocr_text"] = (ocr_text or "").strip()
                # Pré-processa a estrutura da nota (quando houver linhas tipo extrato)
                # para o Excel já sair no formato detalhado imediatamente.
                struct_df = _extract_struct_rows_from_record(rec)
                if not struct_df.empty:
                    rec["_excel_struct_rows"] = struct_df.to_dict(orient="records")
                results.append(rec)
                score = (rec.get("score_revisao") or "").strip().lower()
                if score == "revisar":
                    metrics["revisar"] = metrics.get("revisar", 0) + 1
                    _append_log_duvida("revisar", name, "Nota marcada para revisar (falta dado ou incoerência)", rec.get("numero_nf"))
                elif score == "verificar":
                    metrics["verificar"] = metrics.get("verificar", 0) + 1
                    _append_log_duvida("verificar", name, "Nota com dúvida (verificar)", rec.get("numero_nf"))
            elif ok and dados is None:
                st.session_state.processing[i]["status"] = "Já incluído (não duplicado)"
                ja_incluidos.append(name)
            else:
                st.session_state.processing[i]["status"] = "Erro"
                metrics["erros_total"] = metrics.get("erros_total", 0) + 1
                _append_log_duvida("erro", name, "Processamento falhou", None)
        except Exception as e:
            err_full = str(e).strip()
            err_msg = err_full[:320] if len(err_full) > 320 else err_full
            st.session_state.processing[i]["status"] = f"Erro: {err_msg}"
            if "Nenhuma LLM disponível" in err_full or "LLM" in err_full:
                cause = getattr(e, "__cause__", None)
                st.session_state["llm_last_error"] = {
                    "msg": err_full,
                    "detail": str(cause) if cause else None,
                }
            metrics["erros_total"] = metrics.get("erros_total", 0) + 1
            _append_log_duvida("erro", name, err_msg, None)

    st.session_state.processing[i].pop("bytes", None)
    if ja_incluidos:
        st.session_state.ultimos_ja_incluidos = ja_incluidos
    st.session_state.results = results
    st.session_state.metrics.update(metrics)

    if len(to_run) == 1:
        for p in st.session_state.processing:
            p.pop("bytes", None)
        finished = len([p for p in st.session_state.processing if p.get("status") == "Finalizado"])
        st.session_state.metrics["files_total"] = st.session_state.metrics.get("files_total", 0) + finished
        if results:
            times = [r.get("_elapsed_sec", 0) for r in results if "_elapsed_sec" in r]
            if times:
                st.session_state.metrics["avg_ocr_sec"] = round(sum(times) / len(times), 1)
        total = len(st.session_state.processing)
        ok_count = len([p for p in st.session_state.processing if p.get("status") == "Finalizado"])
        st.session_state.metrics["success_rate"] = round(100 * ok_count / total, 0) if total else 100
        st.session_state.ultima_execucao = datetime.now().strftime("hoje %H:%M")
        _save_state()
        return False
    return True


def results_to_dataframe(results=None):
    if results is None:
        results = st.session_state.get("results", [])
    if not results:
        return pd.DataFrame()
    rows = []
    for r in results:
        rows.append({
            "Empresa": r.get("razao_social_emitente", ""),
            "CNPJ": r.get("cnpj_emitente", ""),
            "Data": r.get("data_emissao", ""),
            "Valor": f"{r.get('valor_total', '')} {r.get('moeda', 'BRL')}".strip(),
            "Rubrica": r.get("rubrica", ""),
            "Pesquisador": r.get("nome_comprador", ""),
            "Produto": (r.get("itens", "") or r.get("discriminacao", ""))[:120],
        })
    return pd.DataFrame(rows)


def _build_full_csv(results=None):
    results = results or st.session_state.get("results", [])
    if not results:
        return ""
    header = list(results[0].keys())
    out = [";".join(header)]
    for r in results:
        out.append(";".join(str(r.get(k, "")) for k in header))
    return "\n".join(out)


def _build_ocr_txt(results=None):
    results = results or st.session_state.get("results", [])
    if not results:
        return ""
    blocks = []
    for i, r in enumerate(results, start=1):
        file_name = (r.get("discriminacao") or f"arquivo_{i}").strip()
        ocr_text = (r.get("_ocr_text") or "").strip()
        if not ocr_text:
            continue
        blocks.append(
            f"[ARQUIVO {i}] {file_name}\n"
            + "-" * 72
            + f"\n{ocr_text}\n"
        )
    return "\n".join(blocks).strip()


def _to_float_br(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    txt = str(value).strip()
    if not txt:
        return 0.0
    txt = txt.replace("R$", "").replace(" ", "")
    # Trata formatos pt-BR e en-US.
    if "," in txt and "." in txt:
        txt = txt.replace(".", "").replace(",", ".")
    elif "," in txt:
        txt = txt.replace(",", ".")
    try:
        return float(txt)
    except Exception:
        return 0.0


def _build_excel_despesas_refeicoes(results=None):
    results = results or st.session_state.get("results", [])
    if not results:
        return b""
    rows = []
    for r in results:
        valor = _to_float_br(r.get("valor_total", ""))
        produto = (r.get("itens", "") or r.get("discriminacao", "") or "").strip()
        rows.append(
            {
                "Data": r.get("data_emissao", ""),
                "Pesquisador": r.get("nome_comprador", ""),
                "Empresa": r.get("razao_social_emitente", ""),
                "CNPJ": r.get("cnpj_emitente", ""),
                "Descrição": r.get("discriminacao", ""),
                "Rubrica": r.get("rubrica", ""),
                "Produto": produto,
                "Valor total": valor,
                "Moeda": r.get("moeda", "BRL"),
            }
        )
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Para uma nota única, prioriza estrutura detalhada em UMA aba só.
        if len(results) == 1:
            det_df = _extract_struct_rows_from_record(results[0])
            # Só considera estrutura quando realmente extraiu valor/total.
            det_ok = (
                not det_df.empty
                and "VALOR" in det_df.columns
                and "TOTAL" in det_df.columns
                and det_df[["VALOR", "TOTAL"]].astype(str).apply(lambda s: s.str.strip()).ne("").any().any()
            )
            if det_ok:
                det_df.to_excel(writer, index=False, sheet_name="Prestacao_Contas")
                _style_structured_worksheet(writer.book["Prestacao_Contas"])
            else:
                pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Prestacao_Contas")
        else:
            pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Prestacao_Contas")
    return buffer.getvalue()


def _style_structured_worksheet(ws):
    """Aplica visual semelhante ao extrato da nota."""
    header_fill = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Corpo
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Alinhamentos por coluna
    col_idx = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}
    for name in ("DATA",):
        idx = col_idx.get(name)
        if idx:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx).alignment = Alignment(horizontal="center", vertical="center")
    for name in ("VALOR", "TOTAL"):
        idx = col_idx.get(name)
        if idx:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx).alignment = Alignment(horizontal="right", vertical="center")

    # Larguras aproximadas
    widths = {"A": 12, "B": 54, "C": 14, "D": 14}
    for col_letter, width in widths.items():
        if ws.column_dimensions.get(col_letter) is not None:
            ws.column_dimensions[col_letter].width = width


def _extract_ocr_table_rows(ocr_text):
    text = str(ocr_text or "")
    if not text.strip():
        return pd.DataFrame()

    # Formato em uma linha (quando OCR já junta tudo).
    full_line_pattern = re.compile(
        r"(?P<data>\d{2}/\d{2}/\d{4})[:\s]+"
        r"(?P<hora>\d{2}:\d{2})\s+"
        r"(?P<orig>\d+)"
        r"(?:\s+(?P<comanda>\d+))?\s+"
        r"(?P<produto>.+?)\s+"
        r"(?P<qtde>\d+,\d{3})\s+R\$\s*"
        r"(?P<valor>[0-9\.,]+)\s+R\$\s*"
        r"(?P<total>[0-9\.,]+)\s*$",
        flags=re.IGNORECASE,
    )
    # Formato da primeira linha (data/hora/comanda/produto) sem valores.
    head_pattern = re.compile(
        r"^(?P<data>\d{2}/\d{2}/\d{4})[:\s]+"
        r"(?P<hora>\d{2}:\d{2})\s+"
        r"(?:(?P<orig>\d+)\s+)?"
        r"(?:(?P<comanda>\d+)\s+)?"
        r"(?P<produto>.+?)$",
        flags=re.IGNORECASE,
    )
    def _normalize_money_text(v):
        s = str(v or "").replace("\t", "").replace(" ", "")
        if not s:
            return ""
        s = s.replace(".", ",")
        if "," not in s and s.isdigit():
            return f"{s},00"
        if "," in s:
            left, right = s.split(",", 1)
            if len(right) == 1:
                right = right + "0"
            elif len(right) > 2:
                right = right[:2]
            return f"{left},{right}"
        return s

    # Formato da segunda linha (qtde + valor + total).
    money_pattern = re.compile(
        r"^(?P<qtde>\d+[,\.\s]?\d{3})\s+R\$\s*(?P<valor>[0-9\.,\s]+)(?:\s+R\$\s*(?P<total>[0-9\.,\s]+))?\s*$",
        flags=re.IGNORECASE,
    )
    total_only_pattern = re.compile(r"^R\$\s*(?P<total>[0-9\.,\s]+)\s*$", flags=re.IGNORECASE)
    total_line_pattern = re.compile(
        r"^(?P<label>TOTAL\s+[A-ZÁÀÃÂÉÊÍÓÔÕÚÇ\s]+)\s+(?P<qtde>\d+,\d{3})\s+R\$\s*(?P<total>[0-9\.,]+)\s*$",
        flags=re.IGNORECASE,
    )

    lines = [" ".join(raw.strip().split()) for raw in text.splitlines() if raw and raw.strip()]
    rows = []
    i = 0
    while i < len(lines):
        line = lines[i]

        # 1) Tenta formato completo em uma linha.
        m_full = full_line_pattern.search(line)
        if m_full:
            rows.append(
                {
                    "DATA": m_full.group("data"),
                    "HORA": m_full.group("hora"),
                    "ORIG": m_full.group("orig") or "",
                    "COMANDA": m_full.group("comanda") or "",
                    "PRODUTO": m_full.group("produto").strip(),
                    "QTDE": _normalize_money_text(m_full.group("qtde")),
                    "VALOR": _normalize_money_text(m_full.group("valor")),
                    "TOTAL": _normalize_money_text(m_full.group("total")),
                }
            )
            i += 1
            continue

        # 1.5) Linha de total (ex.: TOTAL DIARIA ...).
        m_total = total_line_pattern.search(line)
        if m_total:
            rows.append(
                {
                    "DATA": "",
                    "HORA": "",
                    "ORIG": "",
                    "COMANDA": "",
                    "PRODUTO": m_total.group("label").strip(),
                    "QTDE": _normalize_money_text(m_total.group("qtde")),
                    "VALOR": "",
                    "TOTAL": _normalize_money_text(m_total.group("total")),
                }
            )
            i += 1
            continue

        # 2) Tenta formato quebrado em 2-4 linhas (cabeçalho -> produto -> valores -> total).
        m_head = head_pattern.search(line)
        if m_head:
            produto = (m_head.group("produto") or "").strip()
            cursor = i + 1
            consumed = 1

            # Em muitos OCRs, a linha de produto vem separada da linha com data/hora/orig.
            if (not produto) and cursor < len(lines):
                candidate = lines[cursor]
                if (
                    not head_pattern.search(candidate)
                    and not money_pattern.search(candidate)
                    and not total_only_pattern.search(candidate)
                    and "DATA HORA ORIG COMANDA PRODUTO" not in candidate.upper()
                ):
                    produto = candidate.strip()
                    cursor += 1
                    consumed += 1

            qtde = valor = total = ""
            if cursor < len(lines):
                m_money = money_pattern.search(lines[cursor])
                if m_money:
                    qtde = _normalize_money_text(m_money.group("qtde"))
                    valor = _normalize_money_text(m_money.group("valor"))
                    total = _normalize_money_text(m_money.group("total"))
                    cursor += 1
                    consumed += 1
                    # Se total não veio na mesma linha, tenta próxima linha "R$ xx,xx".
                    if (not total) and cursor < len(lines):
                        m_total_only = total_only_pattern.search(lines[cursor])
                        if m_total_only:
                            total = _normalize_money_text(m_total_only.group("total"))
                            consumed += 1

            # Ignora cabeçalhos/linhas inválidas.
            if produto and not produto.upper().startswith("TOTAL ") and (valor or total):
                rows.append(
                    {
                        "DATA": m_head.group("data"),
                        "HORA": m_head.group("hora"),
                        "ORIG": m_head.group("orig") or "",
                        "COMANDA": m_head.group("comanda") or "",
                        "PRODUTO": produto,
                        "QTDE": qtde,
                        "VALOR": valor,
                        "TOTAL": total,
                    }
                )
                i += consumed
                continue

        i += 1

    if rows:
        return pd.DataFrame(rows)

    # 3) Fallback mais tolerante: OCR pode quebrar campos em blocos irregulares.
    # Ex.: linha com DATA/HORA/PRODUTO e outra linha com QTDE/VALOR/TOTAL.
    head_loose_pattern = re.compile(
        r"^(?P<data>\d{2}/\d{2}/\d{4})[:\s]+(?P<hora>\d{2}:\d{2})\s*(?P<rest>.*)$",
        flags=re.IGNORECASE,
    )
    money_loose_pattern = re.compile(
        r"(?P<qtde>\d+[,\.\s]?\d{3}).*?(?P<valor>\d{1,3}(?:[\.\s]\d{3})*[,\.\s]\d{1,2})(?:.*?(?P<total>\d{1,3}(?:[\.\s]\d{3})*[,\.\s]\d{1,2}))?",
        flags=re.IGNORECASE,
    )
    total_only_loose_pattern = re.compile(r"^\s*R\$\s*(?P<total>\d{1,3}(?:[\.\s]\d{3})*[,\.\s]\d{1,2})\s*$", flags=re.IGNORECASE)

    i = 0
    while i < len(lines):
        m_head = head_loose_pattern.search(lines[i])
        if not m_head:
            i += 1
            continue

        data = m_head.group("data")
        hora = m_head.group("hora")
        rest = (m_head.group("rest") or "").strip()

        orig = ""
        comanda = ""
        produto = ""

        # Tenta capturar ORIG/COMANDA no começo do "rest".
        m_ids = re.match(r"^(?P<orig>\d+)(?:\s+(?P<comanda>\d+))?\s*(?P<produto>.*)$", rest)
        if m_ids:
            orig = (m_ids.group("orig") or "").strip()
            comanda = (m_ids.group("comanda") or "").strip()
            produto = (m_ids.group("produto") or "").strip()
        else:
            produto = rest

        # Busca linha de valores no bloco atual + próximas 1..4 linhas.
        qtde = valor = total = ""
        for j in range(i, min(i + 5, len(lines))):
            mm = money_loose_pattern.search(lines[j])
            if mm:
                qtde = _normalize_money_text(mm.group("qtde"))
                valor = _normalize_money_text(mm.group("valor"))
                total = _normalize_money_text(mm.group("total"))
                i = j  # avança cursor até linha de valores
                if (not total) and j + 1 < len(lines):
                    mt = total_only_loose_pattern.search(lines[j + 1])
                    if mt:
                        total = _normalize_money_text(mt.group("total"))
                        i = j + 1
                break

        if produto and (valor or total):
            produto_up = produto.upper()
            if not produto_up.startswith("TOTAL "):
                rows.append(
                    {
                        "DATA": data,
                        "HORA": hora,
                        "ORIG": orig,
                        "COMANDA": comanda,
                        "PRODUTO": produto,
                        "QTDE": qtde,
                        "VALOR": valor,
                        "TOTAL": total,
                    }
                )
        i += 1

    return pd.DataFrame(rows)


def _extract_hotel_statement_rows(ocr_text):
    """Parser dedicado para layout 'Nota de Hospedagem' (blocos de 3-4 linhas)."""
    text = str(ocr_text or "")
    if not text.strip():
        return pd.DataFrame()

    def _normalize_money_text(v):
        s = str(v or "").replace("\t", "").replace(" ", "")
        if not s:
            return ""
        s = s.replace(".", ",")
        if "," not in s and s.isdigit():
            return f"{s},00"
        if "," in s:
            left, right = s.split(",", 1)
            if len(right) == 1:
                right = right + "0"
            elif len(right) > 2:
                right = right[:2]
            return f"{left},{right}"
        return s

    lines = [" ".join(raw.strip().split()) for raw in text.splitlines() if raw and raw.strip()]
    if not lines:
        return pd.DataFrame()

    head_pattern = re.compile(
        r"^(?P<data>\d{2}/\d{2}/\d{4})[:\s]+"
        r"(?P<hora>\d{2}:\d{2})\s+"
        r"(?P<orig>[A-Z]?\d+)"
        r"(?:\s+(?P<comanda>[A-Z]?\d+))?"
        r"(?:\s+(?P<produto_inline>.+))?\s*$",
        flags=re.IGNORECASE,
    )
    # Alguns blocos (ex.: restaurante) vêm sem ORIG/COMANDA.
    head_pattern_no_orig = re.compile(
        r"^(?P<data>\d{2}/\d{2}/\d{4})[:\s]+"
        r"(?P<hora>\d{2}:\d{2})"
        r"(?:\s+(?P<produto_inline>.+))?\s*$",
        flags=re.IGNORECASE,
    )
    qty_val_pattern = re.compile(
        r"^(?P<qtde>\d+[,\.\s]?\d{3})\s+R\$\s*(?P<valor>[0-9\.,\s]+)\s*$",
        flags=re.IGNORECASE,
    )
    qty_only_pattern = re.compile(r"^(?P<qtde>\d+[,\.\s]?\d{3})\s*$", flags=re.IGNORECASE)
    money_only_pattern = re.compile(r"^R\$\s*(?P<money>[0-9\.,\s]+)\s*$", flags=re.IGNORECASE)
    money_amount_only_pattern = re.compile(r"^(?P<money>[0-9\.,\s]+)\s*$", flags=re.IGNORECASE)
    money_symbol_only_pattern = re.compile(r"^R\$\s*$", flags=re.IGNORECASE)
    total_only_pattern = re.compile(r"^R\$\s*(?P<total>[0-9\.,\s]+)\s*$", flags=re.IGNORECASE)
    total_label_pattern = re.compile(r"^(?P<label>TOTAL\s+[A-ZÁÀÃÂÉÊÍÓÔÕÚÇ\s]+)$", flags=re.IGNORECASE)

    rows = []
    i = 0
    while i < len(lines):
        line = lines[i]
        line_up = line.upper()

        # Ignora cabeçalho da tabela.
        if "DATA HORA ORIG COMANDA PRODUTO" in line_up or line_up in {"QTDE", "VALOR", "TOTAL"}:
            i += 1
            continue

        # Bloco TOTAL: ignora (não deve entrar como PRODUTO).
        if total_label_pattern.match(line):
            i += 1
            continue

        # Bloco item: data/hora/orig(+comanda) [produto inline opcional]
        # -> produto (se necessário) -> qtde+valor -> total
        m_head = head_pattern.match(line)
        head_has_orig = True
        if not m_head:
            m_head = head_pattern_no_orig.match(line)
            head_has_orig = False
        if m_head and i + 2 < len(lines):
            produto_inline = (m_head.group("produto_inline") or "").strip()
            consumed = 1
            if produto_inline:
                produto = produto_inline
                qv_idx = i + 1
            else:
                produto = lines[i + 1].strip()
                qv_idx = i + 2
                consumed = 2

            # Se "produto" veio só com código numérico, tenta a próxima linha textual.
            if re.fullmatch(r"\d+", produto or "") and qv_idx < len(lines):
                alt_prod_idx = qv_idx
                if alt_prod_idx < len(lines):
                    alt_prod = lines[alt_prod_idx].strip()
                    if alt_prod and not re.fullmatch(r"[\d\.,:]+", alt_prod):
                        produto = alt_prod
                        qv_idx = min(alt_prod_idx + 1, len(lines) - 1)
                        consumed += 1

            qtde = valor = total = ""
            # Caso A: qtde e valor na mesma linha
            m_qv = qty_val_pattern.match(lines[qv_idx]) if qv_idx < len(lines) else None
            if m_qv:
                qtde = _normalize_money_text(m_qv.group("qtde"))
                valor = _normalize_money_text(m_qv.group("valor"))
                if qv_idx + 1 < len(lines):
                    m_tot = total_only_pattern.match(lines[qv_idx + 1])
                    if m_tot:
                        total = _normalize_money_text(m_tot.group("total"))
                        consumed += 2
                    else:
                        # OCR pode vir só com número do total (sem "R$").
                        m_tot_amount = money_amount_only_pattern.match(lines[qv_idx + 1])
                        if m_tot_amount:
                            total = _normalize_money_text(m_tot_amount.group("money"))
                            consumed += 2
                        else:
                            total = valor
                            consumed += 1
                else:
                    total = valor
                    consumed += 1
            else:
                # Caso B: qtde em uma linha e valor/total em linhas seguintes
                m_q = qty_only_pattern.match(lines[qv_idx]) if qv_idx < len(lines) else None
                if m_q:
                    qtde = _normalize_money_text(m_q.group("qtde"))
                    next_idx = qv_idx + 1

                    # Valor pode vir em: "R$ 55,00" | "R$" + "55,00" | "55,00"
                    if next_idx < len(lines):
                        m_val = money_only_pattern.match(lines[next_idx])
                        if m_val:
                            valor = _normalize_money_text(m_val.group("money"))
                            next_idx += 1
                        elif money_symbol_only_pattern.match(lines[next_idx]) and next_idx + 1 < len(lines):
                            m_val2 = money_amount_only_pattern.match(lines[next_idx + 1])
                            if m_val2:
                                valor = _normalize_money_text(m_val2.group("money"))
                                next_idx += 2
                        else:
                            m_val3 = money_amount_only_pattern.match(lines[next_idx])
                            if m_val3:
                                valor = _normalize_money_text(m_val3.group("money"))
                                next_idx += 1

                    # Total pode vir em: "R$ 55,00" | "55,00"
                    if next_idx < len(lines):
                        m_tot = total_only_pattern.match(lines[next_idx])
                        if m_tot:
                            total = _normalize_money_text(m_tot.group("total"))
                            next_idx += 1
                        else:
                            m_tot2 = money_amount_only_pattern.match(lines[next_idx])
                            if m_tot2:
                                total = _normalize_money_text(m_tot2.group("money"))
                                next_idx += 1

                    if not total:
                        total = valor

                    consumed = max(consumed, next_idx - i)

            if valor or total:
                produto_up = str(produto or "").upper().strip()
                is_origin_like_code = bool(re.fullmatch(r"[A-Z]{0,3}\d{4,}", produto_up))
                next_is_total_label = False
                if i + consumed < len(lines):
                    next_is_total_label = bool(total_label_pattern.match(lines[i + consumed]))

                # Se a próxima linha é TOTAL e qtde não é unitária, normalmente é subtotal do bloco.
                qtde_is_unitary = str(qtde or "").replace(" ", "") in {"1,000", "1.000", "1,00", "1.00"}
                is_subtotal_like = next_is_total_label and (not qtde_is_unitary)

                if produto_up and (not produto_up.startswith("TOTAL ")) and (not is_origin_like_code) and (not is_subtotal_like):
                    rows.append(
                        {
                            "DATA": m_head.group("data"),
                            "PRODUTO": produto,
                            "VALOR": valor,
                            "TOTAL": total,
                        }
                    )
                    i += consumed
                    continue

        i += 1

    return pd.DataFrame(rows)


def _extract_struct_rows_from_record(record):
    """Monta linhas estruturadas para a aba Estrutura_Nota."""
    text = record.get("_ocr_text", "")

    # 1) Parser dedicado para Nota de Hospedagem (mais fiel ao layout).
    df_hotel = _extract_hotel_statement_rows(text)
    # 2) Parser genérico (complementa linhas que o dedicado pode perder).
    df_generic = _extract_ocr_table_rows(text)

    frames = []
    if not df_hotel.empty:
        frames.append(df_hotel)
    if not df_generic.empty:
        frames.append(df_generic)
    if frames:
        merged = pd.concat(frames, ignore_index=True)
        # Não deduplicar aqui: sem HORA, itens legítimos repetidos podem sumir.
        return _sanitize_struct_rows_df(merged)

    # Fallback: quando não há OCR completo, tenta montar lista mínima com produtos.
    raw_itens = record.get("itens")
    itens_list = []
    if isinstance(raw_itens, list):
        itens_list = [str(x).strip() for x in raw_itens if str(x).strip()]
    elif isinstance(raw_itens, str) and raw_itens.strip():
        try:
            parsed = json.loads(raw_itens)
            if isinstance(parsed, list):
                itens_list = [str(x).strip() for x in parsed if str(x).strip()]
        except Exception:
            itens_list = [x.strip() for x in raw_itens.split(";") if x.strip()]
    # Se itens vierem como blocos multiline (data/produto/valor), tenta parsear.
    if itens_list:
        joined = "\n".join(itens_list)
        # Também combina os dois parsers no fallback de itens.
        if joined.strip():
            frames_items = [df for df in (_extract_hotel_statement_rows(joined), _extract_ocr_table_rows(joined)) if not df.empty]
            df_from_items = pd.concat(frames_items, ignore_index=True) if frames_items else pd.DataFrame()
        else:
            df_from_items = pd.DataFrame()
        if not df_from_items.empty:
            return _sanitize_struct_rows_df(df_from_items)
    if not itens_list:
        return pd.DataFrame()
    return _sanitize_struct_rows_df(pd.DataFrame(
        [{"DATA": "", "PRODUTO": it, "VALOR": "", "TOTAL": ""} for it in itens_list]
    ))


def _sanitize_struct_rows_df(df):
    """Garante saída final somente com DATA, PRODUTO, VALOR, TOTAL e sem linhas vazias."""
    if df is None or df.empty:
        return pd.DataFrame(columns=["DATA", "PRODUTO", "VALOR", "TOTAL"])
    out = df.copy()

    # Preenche produto quando parser deixou em branco.
    if "PRODUTO" not in out.columns:
        out["PRODUTO"] = ""
    for alt_col in ("COMANDA", "ORIG", "HORA"):
        if alt_col in out.columns:
            alt = out[alt_col].astype(str).str.strip()
            prod = out["PRODUTO"].astype(str).str.strip()
            # Só usa fallback se parecer texto de item (não só dígitos/horário).
            is_text_like = ~alt.str.fullmatch(r"[\d:.,\s]+")
            use_mask = prod.eq("") & alt.ne("") & is_text_like
            out.loc[use_mask, "PRODUTO"] = alt[use_mask]

    # Mantém apenas as colunas desejadas.
    for col in ("DATA", "VALOR", "TOTAL"):
        if col not in out.columns:
            out[col] = ""
    out = out[["DATA", "PRODUTO", "VALOR", "TOTAL"]]

    # Remove linhas sem informação útil.
    keep = (
        out["PRODUTO"].astype(str).str.strip().ne("")
        | out["VALOR"].astype(str).str.strip().ne("")
        | out["TOTAL"].astype(str).str.strip().ne("")
    )
    out = out[keep].copy()

    # Descarta linhas em que "produto" parece código (numérico ou alfanumérico tipo E000277).
    prod_clean = out["PRODUTO"].astype(str).str.strip().str.upper()
    is_numeric_code = prod_clean.str.fullmatch(r"\d+")
    is_alnum_code = prod_clean.str.fullmatch(r"[A-Z]{0,3}\d{4,}")
    out = out[~(is_numeric_code | is_alnum_code)]

    # Não deduplica: sem HORA/COMANDA, repetições podem ser itens válidos.
    return out.reset_index(drop=True)


def _build_csv_for_record(record):
    """Gera CSV individual; prioriza estrutura extraída do OCR."""
    def _is_good_struct_df(df):
        if df is None or df.empty:
            return False
        needed = {"PRODUTO", "VALOR", "TOTAL"}
        if not needed.issubset(set(df.columns)):
            return False
        prod_ok = df["PRODUTO"].astype(str).str.strip().ne("").sum()
        money_ok = (
            df["VALOR"].astype(str).str.strip().ne("").sum()
            + df["TOTAL"].astype(str).str.strip().ne("").sum()
        )
        # Só aceita estrutura quando há produtos + valores de fato.
        return prod_ok > 0 and money_ok > 0

    df = _extract_struct_rows_from_record(record)

    # Se cache vier ruim (ex.: só PRODUTO), reparseia do OCR em tempo real.
    if not _is_good_struct_df(df):
        reparsed = _extract_ocr_table_rows(record.get("_ocr_text", ""))
        if _is_good_struct_df(reparsed):
            df = reparsed

    if _is_good_struct_df(df):
        return df.to_csv(index=False, sep=";", encoding="utf-8")

    # Evita baixar CSV "quebrado" com colunas vazias; cai no resumo.
    return _build_full_csv([record])


def _excel_export_filename(results=None):
    results = results or st.session_state.get("results", [])
    if not results:
        return "notas_fiscais.xlsx"
    originals = []
    for r in results:
        nm = str(r.get("discriminacao", "") or "").strip()
        if nm:
            originals.append(Path(nm).stem)
    originals = [o for o in originals if o]
    if not originals:
        return "notas_fiscais.xlsx"
    base = originals[0].strip()
    if not base:
        base = "nota_fiscal"
    if len(originals) == 1:
        return f"{base}.xlsx"
    return f"{base}_lote.xlsx"


def _csv_export_filename(results=None):
    results = results or st.session_state.get("results", [])
    if not results:
        return "notas_fiscais.csv"
    originals = []
    for r in results:
        nm = str(r.get("discriminacao", "") or "").strip()
        if nm:
            originals.append(Path(nm).stem)
    originals = [o for o in originals if o]
    if not originals:
        return "notas_fiscais.csv"
    base = originals[0].strip() or "nota_fiscal"
    if len(originals) == 1:
        return f"{base}.csv"
    return f"{base}_lote.csv"


def _norm_filename_key(name):
    name = str(name or "").strip().lower()
    # chave robusta para casar nomes com diferenças de espaços/pontuação
    return "".join(ch for ch in name if ch.isalnum())


def _render_processed_exports(records, key_prefix="processed"):
    if not records:
        st.info("Nenhum arquivo processado ainda.")
        return
    st.markdown("**⬇️ Exportar por arquivo (individual)**")
    for i, rec in enumerate(records):
        original_name = str(rec.get("discriminacao", "") or f"nota_{i+1}").strip()
        base = Path(original_name).stem or f"nota_{i+1}"
        one_csv = _build_csv_for_record(rec)
        one_xlsx = _build_excel_despesas_refeicoes([rec])
        one_txt = (rec.get("_ocr_text") or "").strip()
        st.caption(f"📄 **{original_name}**")
        col_csv_one, col_xlsx_one, col_txt_one = st.columns(3)
        with col_csv_one:
            st.download_button(
                "CSV",
                data=one_csv,
                file_name=f"{base}.csv",
                mime="text/csv",
                use_container_width=True,
                key=f"{key_prefix}_dl_csv_single_{i}",
            )
        with col_xlsx_one:
            if one_xlsx:
                st.download_button(
                    "Excel",
                    data=one_xlsx,
                    file_name=f"{base}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"{key_prefix}_dl_xlsx_single_{i}",
                )
        with col_txt_one:
            if one_txt:
                st.download_button(
                    "TXT",
                    data=one_txt,
                    file_name=f"{base}.txt",
                    mime="text/plain",
                    use_container_width=True,
                    key=f"{key_prefix}_dl_txt_single_{i}",
                )


def _reclassify_processed_results():
    """Reclassifica rubrica e atualiza estrutura Excel dos registros salvos."""
    try:
        import nf_ocr
    except Exception as e:
        return 0, f"Falha ao importar nf_ocr: {e}"
    results = list(st.session_state.get("results", []))
    if not results:
        return 0, None
    changed = 0
    excel_updated = 0
    updated = []
    for rec in results:
        before = str(rec.get("rubrica") or "").strip().lower()
        text = str(rec.get("_ocr_text") or "")
        if hasattr(nf_ocr, "reclassificar_rubrica_processada"):
            new_rec = nf_ocr.reclassificar_rubrica_processada(rec, text)
        else:
            # Fallback para ambientes com cache/stale import do nf_ocr.
            text_low = (text or "").lower()
            itens = rec.get("itens")
            if isinstance(itens, list):
                text_low += " " + " ".join(str(x).lower() for x in itens)
            elif isinstance(itens, str):
                text_low += " " + itens.lower()
            text_low += " " + str(rec.get("discriminacao") or "").lower()
            text_norm = "".join(
                ch for ch in unicodedata.normalize("NFD", text_low) if unicodedata.category(ch) != "Mn"
            )
            sinais_hosp = [
                "hospedagem", "hotel", "diaria", "check-in", "check-out",
                "hospede", "total diaria", "apto",
            ]
            new_rec = dict(rec)
            if any(s in text_norm for s in sinais_hosp):
                new_rec["rubrica"] = "viagem"
        after = str(new_rec.get("rubrica") or "").strip().lower()
        if after != before:
            changed += 1
        struct_df = _extract_struct_rows_from_record(new_rec)
        if not struct_df.empty:
            new_rec["_excel_struct_rows"] = struct_df.to_dict(orient="records")
            excel_updated += 1
        updated.append(new_rec)
    st.session_state.results = updated
    _save_state()
    return changed, excel_updated, None


def page_inicio():
    metrics = st.session_state.get("metrics", {})
    processing = st.session_state.get("processing", [])

    col_main, col_right = st.columns([4, 1])
    with col_main:
        st.markdown('<div class="nf-card-panel"><h4>📊 Métricas do sistema</h4>', unsafe_allow_html=True)
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.markdown(f'<div class="nf-metric"><div class="nf-metric-value">{metrics.get("files_total", 0)}</div><div class="nf-metric-label">Arquivos processados</div></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="nf-metric"><div class="nf-metric-value">{metrics.get("avg_ocr_sec", 0)}s</div><div class="nf-metric-label">Tempo médio OCR</div></div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div class="nf-metric"><div class="nf-metric-value">{metrics.get("success_rate", 100)}%</div><div class="nf-metric-label">Taxa de sucesso</div></div>', unsafe_allow_html=True)
        with c4:
            st.markdown(f'<div class="nf-metric"><div class="nf-metric-value">{metrics.get("fields_detected", 12)}</div><div class="nf-metric-label">Campos detectados</div></div>', unsafe_allow_html=True)
        with c5:
            st.markdown(f'<div class="nf-metric"><div class="nf-metric-value">{metrics.get("erros_total", 0)}</div><div class="nf-metric-label">Erros</div></div>', unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown('<div class="nf-card-panel"><h4>☁️ Upload de notas fiscais</h4>', unsafe_allow_html=True)
        st.markdown(
            '<div class="nf-upload-hero"><div class="cloud">☁️</div><p style="margin:0;color:#475569;">Arraste arquivos de notas fiscais aqui</p></div>',
            unsafe_allow_html=True,
        )
        uploaded = st.file_uploader(
            "ou Selecionar arquivos",
            type=["pdf", "png", "jpg", "jpeg"],
            accept_multiple_files=True,
            key="uploader_main",
            label_visibility="visible",
        )
        st.caption("Formatos aceitos: PDF, PNG, JPG, JPEG · Máximo 200 MB por arquivo")
        st.markdown("</div>", unsafe_allow_html=True)

        if uploaded:
            max_mb = 200
            valid = [f for f in uploaded if f.size / (1024 * 1024) <= max_mb]
            for f in uploaded:
                if f.size / (1024 * 1024) > max_mb:
                    st.warning(f"**{f.name}** excede {max_mb} MB e foi ignorado.")
            if valid:
                clicou = st.button("► Iniciar processamento", type="primary", use_container_width=True)
                if clicou:
                    _load_secrets_to_env()
                    _load_gcp_secret_manager_to_env()
                    groq = os.environ.get("GROQ_API_KEY", "")
                    hf = os.environ.get("HF_TOKEN", "")
                    if not groq and not hf:
                        gcp_err = (os.environ.get("GCP_SECRET_MANAGER_LAST_ERROR") or "").strip()
                        gcp_proj = (os.environ.get("GCP_SECRET_MANAGER_PROJECT_USED") or "").strip()
                        gac = (os.environ.get("GOOGLE_APPLICATION_CREDENTIALS") or "").strip()
                        st.error(
                            "⚠️ **Chave de API não definida.**\n\n"
                            "O app tentou carregar automaticamente de: **Streamlit Secrets**, **variáveis de ambiente** e **Google Secret Manager**.\n\n"
                            "**Google Secret Manager:** o segredo deve se chamar **`GROQ_API_KEY`** no **mesmo projeto** onde está a service account (ou defina `GOOGLE_CLOUD_PROJECT`). "
                            "A conta precisa da role **Secret Manager Secret Accessor** nesse segredo.\n\n"
                            "**No seu PC (fallback):** crie/edite **.env** com `GROQ_API_KEY=...` ou `GOOGLE_APPLICATION_CREDENTIALS=C:\\\\caminho\\\\chave.json`.\n"
                            "**Windows:** depois de `export`/`set`, **feche e abra o terminal** e rode o Streamlit de novo (IDEs às vezes não herdam o env).\n"
                            "**Na nuvem (fallback):** use **Settings → Secrets**."
                        )
                        if gcp_proj:
                            st.caption(f"Projeto GCP usado no Secret Manager: **{gcp_proj}**")
                        if gac:
                            _exists = Path(gac).is_file()
                            st.caption(f"`GOOGLE_APPLICATION_CREDENTIALS` = `{gac}` — arquivo existe: **{_exists}**")
                        if gcp_err:
                            st.caption(f"Diagnóstico GCP: {gcp_err}")
                        return
                    if groq:
                        os.environ["GROQ_API_KEY"] = groq
                    if hf:
                        os.environ["HF_TOKEN"] = hf
                    with st.spinner("Enviando..."):
                        st.session_state.processing = [
                            {"name": f.name, "status": "Enviado", "progress": 0, "bytes": f.getvalue()}
                            for f in valid
                        ]
                    st.rerun()

    with col_right:
        st.markdown('<div class="nf-card-panel"><h4>📋 Atividade recente</h4>', unsafe_allow_html=True)
        recent = (processing[-6:])[::-1]
        if recent:
            for item in recent:
                name = item.get("name", "?")
                status = item.get("status", "Enviado")
                st.caption(f"📄 **{name}** — {status}")
            st.caption("_Ver todas >_")
        else:
            st.caption("_Nenhum arquivo processado nesta sessão._")
        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown('<div class="nf-card-panel"><h4>📈 Resumo do sistema</h4><p style="margin:0 0 0.5rem 0;font-size:0.85rem;color:#64748B;">Hoje</p>', unsafe_allow_html=True)
        m = st.session_state.get("metrics", {})
        st.caption(f"**{m.get('files_total', 0)}** Arquivos processados")
        st.caption(f"**{m.get('avg_ocr_sec', 0)}s** Tempo médio OCR")
        st.caption(f"**{m.get('success_rate', 100)}%** Taxa de sucesso")
        st.caption(f"**{m.get('fields_detected', 12)}** Campos detectados")
        st.markdown("</div>", unsafe_allow_html=True)

    with col_main:
        # Erro do DocTR fica guardado para não sumir após o rerun
        if st.session_state.get("doctr_last_error"):
            err_info = st.session_state.doctr_last_error
            st.markdown("---")
            st.error(f"**Último erro ao carregar DocTR:** {err_info.get('msg', '')}")
            with st.expander("Traceback completo (copie e envie para diagnóstico)"):
                st.code(err_info.get("tb", ""), language="text")
            if st.button("Limpar esta mensagem"):
                del st.session_state["doctr_last_error"]
                st.rerun()
        if st.session_state.get("llm_last_error"):
            err_info = st.session_state.llm_last_error
            msg = err_info.get("msg", "")
            st.markdown("---")
            st.error("**Erro na LLM (extração com IA)** — veja o detalhe abaixo.")
            if "Invalid API Key" in msg or "invalid_api_key" in msg:
                st.warning(
                    "**Chave da Groq inválida (401).** A `GROQ_API_KEY` está incorreta (em Secret Manager, Secrets do Streamlit ou env). "
                    "**Faça assim:** 1) Abra [console.groq.com/keys](https://console.groq.com/keys) e crie ou copie uma chave (começa com `gsk_`). "
                    "2) Atualize o valor de `GROQ_API_KEY` no **Google Secret Manager** (ou em Secrets do app) — **sem espaços** extras e com a chave completa. "
                    "3) Guarde, espere ~1 min e faça **Reboot** do app."
                )
            elif "Invalid username or password" in msg or ("401" in msg and "huggingface" in msg.lower()):
                st.warning(
                    "**Token do Hugging Face inválido (401).** O erro vem da API do Hugging Face, não da Groq. "
                    "**Soluções:** 1) Use só Groq: em **Secrets** coloque apenas `GROQ_API_KEY` (chave em [console.groq.com](https://console.groq.com/keys)) e remova ou apague o valor de `HF_TOKEN`. "
                    "2) Ou corrija o token HF em [huggingface.co/settings/tokens](https://huggingface.co/settings/tokens) (crie um token novo e atualize nos Secrets)."
                )
            st.code(msg, language="text")
            if err_info.get("detail"):
                st.caption(f"Causa: {err_info['detail']}")
            if st.button("Limpar esta mensagem", key="clear_llm_err"):
                del st.session_state["llm_last_error"]
                st.rerun()
        if processing:
            st.markdown("---")
            st.markdown("#### 2️⃣ Processamento")
            has_pending = any(p.get("bytes") for p in processing)
            if has_pending:
                n_total = len(processing)
                pending_count = sum(1 for p in processing if p.get("bytes") and p.get("status") not in ("Finalizado", "Erro"))
                done_so_far = n_total - pending_count
                pct = (done_so_far * 100) // n_total if n_total else 0
                progress_placeholder = st.empty()
                # Evita "0%" parado: se ainda vai processar, mostra "Processando…"; quando termina o rerun mostra o resultado
                if pending_count > 0 and done_so_far == 0:
                    progress_placeholder.progress(0, text=f"Processando arquivo 1 de {n_total}…")
                else:
                    progress_placeholder.progress(done_so_far / n_total if n_total else 0, text=f"Arquivo {done_so_far + 1} de {n_total} ({pct}%)")
                try:
                    has_more = run_processing(progress_placeholder)
                    if has_more:
                        st.rerun()
                except Exception as e:
                    progress_placeholder.empty()
                    st.error(f"Erro: {e}")
                st.rerun()

            for item in processing:
                name = item.get("name", "?")
                status = item.get("status", "Enviado")
                sc = "finalizado" if status == "Finalizado" else "erro" if "Erro" in str(status) else "enviado" if status == "Enviado" else "ja_incluido" if "Já incluído" in str(status) else "enviado"
                st.markdown(
                    f'<div class="nf-card"><span style="color:#0F172A;font-weight:600;">{name}</span> <span class="nf-badge nf-badge-{sc}">{status}</span></div>',
                    unsafe_allow_html=True,
                )
            if st.session_state.get("ultimos_ja_incluidos"):
                ja_list = st.session_state.ultimos_ja_incluidos
                st.warning(f"**Aviso:** Os seguintes arquivos já constavam na planilha e **não foram duplicados**: {', '.join(ja_list)}.")
                st.session_state.ultimos_ja_incluidos = []
            current_results = st.session_state.get("results", [])
            if current_results:
                result_by_name = {}
                for rec in current_results:
                    nm = str(rec.get("discriminacao", "") or "").strip()
                    if nm:
                        result_by_name.setdefault(nm, rec)
                        result_by_name.setdefault(_norm_filename_key(nm), rec)
                export_records = []
                for i, item in enumerate(processing):
                    status_txt = str(item.get("status", "") or "")
                    if ("Finalizado" not in status_txt) and ("Já incluído" not in status_txt):
                        continue
                    original_name = str(item.get("name", "") or "").strip()
                    rec = result_by_name.get(original_name) or result_by_name.get(_norm_filename_key(original_name))
                    if not rec:
                        continue
                    export_records.append(rec)
                _render_processed_exports(export_records, key_prefix="upload")

        results = st.session_state.get("results", [])
        if results:
            st.markdown("---")
            st.markdown("#### 3️⃣ Resultados")
            st.markdown("Filtre por **rubrica**, **data** ou **pesquisador**. Ao final, **exporte em CSV**.")
            csv_path_cfg = (st.session_state.get("csv_drive_path") or "").strip()
            if not csv_path_cfg:
                st.caption("📁 Os dados são gravados em **nf_dados/nf_extraidas.csv** (pasta do app). Estado e log ficam em **nf_dados/**.")
            df_full = results_to_dataframe(results)

            st.markdown("**🔍 Filtros**")
            rubricas = ["viagem", "participacao_congresso", "material_consumo", "material_permanente", "servico_terceiros"]
            labels = ["Viagem", "Participação em congresso", "Material de consumo", "Material permanente", "Serviço de terceiros"]
            filtro_rubrica = st.radio("Rubrica", options=["Todas"] + labels, key="filtro_rubrica", horizontal=True)
            col_d, col_p = st.columns(2)
            with col_d:
                datas = sorted(set(r.get("data_emissao", "") for r in results if r.get("data_emissao")))
                use_cal = st.checkbox("Usar calendário para data", key="use_cal_data")
                if use_cal:
                    from datetime import date
                    d = st.date_input("📅 Selecione a data", key="filtro_data_cal")
                    filtro_data = d.strftime("%Y-%m-%d") if d else "Todas"
                else:
                    filtro_data = st.selectbox("Data", options=["Todas"] + datas, key="filtro_data")
            with col_p:
                pesquisadores_resultados = set(r.get("nome_comprador", "") for r in results if r.get("nome_comprador"))
                pesquisadores_config = set(st.session_state.get("lista_pesquisadores", []))
                pesquisadores = sorted(pesquisadores_resultados | pesquisadores_config)
                filtro_pesq = st.selectbox("Pesquisador", options=["Todos"] + pesquisadores, key="filtro_pesq")

            filtered = list(results)
            if filtro_rubrica != "Todas":
                rubrica_val = rubricas[labels.index(filtro_rubrica)]
                filtered = [r for r in filtered if (r.get("rubrica") or "").strip() == rubrica_val]
            if filtro_data != "Todas":
                filtered = [r for r in filtered if r.get("data_emissao") == filtro_data]
            if filtro_pesq != "Todos":
                filtered = [r for r in filtered if (r.get("nome_comprador") or "").strip() == filtro_pesq]

            df = results_to_dataframe(filtered)
            # Contorno escuro e minimalista; sem fundo para o texto usar a cor do tema (visível)
            styled = df.style.set_table_styles([
                {"selector": "table", "props": [("border", "1px solid #1e293b"), ("border-collapse", "collapse"), ("border-radius", "8px")]},
                {"selector": "thead th", "props": [("border-bottom", "2px solid #1e293b"), ("border-right", "1px solid #334155")]},
                {"selector": "tbody td", "props": [("border-bottom", "1px solid #e2e8f0"), ("border-right", "1px solid #e2e8f0")]},
                {"selector": "th:first-child, td:first-child", "props": [("border-left", "1px solid #334155")]},
            ])
            st.dataframe(
                styled,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Empresa": st.column_config.TextColumn("Empresa", width="medium"),
                    "CNPJ": st.column_config.TextColumn("CNPJ", width="small"),
                    "Data": st.column_config.TextColumn("Data", width="small"),
                    "Valor": st.column_config.TextColumn("Valor", width="small"),
                    "Rubrica": st.column_config.TextColumn("Rubrica", width="medium"),
                    "Pesquisador": st.column_config.TextColumn("Pesquisador", width="medium"),
                    "Produto": st.column_config.TextColumn("Produto", width="large"),
                },
                key="results_editor",
            )
            st.caption("Os downloads individuais (CSV/Excel/TXT) estão na seção de Processamento.")


def page_revisar():
    st.subheader("⚠️ Área de revisão")
    results = st.session_state.get("results", [])
    revisar = [r for r in results if (r.get("score_revisao") or "").strip().lower() in ("revisar", "verificar")]
    st.markdown("Itens com score **revisar** ou **verificar** (falta dado, incoerência ou dúvida). Revise e corrija na planilha ou no exporto.")
    if not revisar:
        st.info("Nenhum item para revisar no momento. Os que precisarem aparecerão aqui após o processamento.")
        return
    st.markdown(f"**{len(revisar)}** itens para revisar.")
    df = results_to_dataframe(revisar)
    df["Score"] = [r.get("score_revisao", "") for r in revisar]
    df["Arquivo"] = [r.get("discriminacao", "") for r in revisar]
    st.data_editor(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Empresa": st.column_config.TextColumn("Empresa", width="medium"),
            "CNPJ": st.column_config.TextColumn("CNPJ", width="small"),
            "Data": st.column_config.TextColumn("Data", width="small"),
            "Valor": st.column_config.TextColumn("Valor", width="small"),
            "Rubrica": st.column_config.TextColumn("Rubrica", width="medium"),
            "Pesquisador": st.column_config.TextColumn("Pesquisador", width="medium"),
            "Produto": st.column_config.TextColumn("Produto", width="large"),
            "Score": st.column_config.TextColumn("Score", width="small"),
            "Arquivo": st.column_config.TextColumn("Arquivo", width="medium"),
        },
        key="revisar_editor",
    )
    csv_revisar = _build_full_csv(revisar)
    st.download_button("📥 Exportar lista para revisão (CSV)", data=csv_revisar, file_name="notas_para_revisar.csv", mime="text/csv", use_container_width=True, key="dl_revisar")


def page_processados():
    st.subheader("📂 Arquivos processados")
    results = st.session_state.get("results", [])
    if not results:
        st.info("Ainda não há arquivos processados para listar.")
        return
    if st.button("🔄 Atualizar rubricas dos processados", use_container_width=True, key="btn_reclass_processados"):
        changed, excel_updated, err = _reclassify_processed_results()
        if err:
            st.error(err)
        elif changed > 0 or excel_updated > 0:
            st.success(f"Rubrica atualizada em {changed} registro(s) e estrutura Excel atualizada em {excel_updated} registro(s).")
        else:
            st.info("Nenhuma rubrica precisou ser alterada.")
        results = st.session_state.get("results", [])
    st.caption(f"{len(results)} registro(s) processado(s) disponível(is) para download individual.")
    _render_processed_exports(results, key_prefix="page_processados")


def page_configuracoes():
    st.subheader("⚙️ Configurações")
    st.markdown("**📁 Planilha no Drive (ou pasta local)**")
    st.caption("**Por padrão** os dados são salvos em **nf_dados/** (CSV, estado e log). Deixe vazio para usar esse padrão. Se quiser outro local (ex.: pasta do Google Drive), informe o caminho completo do arquivo CSV.")
    default_hint = str(_default_csv_path()) if _default_csv_path() else "nf_dados/nf_extraidas.csv"
    csv_path = st.text_input(
        "Caminho do arquivo CSV (vazio = padrão)",
        value=st.session_state.get("csv_drive_path", "") or "",
        placeholder=default_hint,
        key="input_csv_drive_path",
    )
    if csv_path != st.session_state.get("csv_drive_path", ""):
        st.session_state.csv_drive_path = (csv_path or "").strip()
        _save_state()
    st.markdown("---")
    st.markdown("**👤 Pesquisadores** — os nomes abaixo aparecem no filtro e são enviados para a **IA** reconhecer melhor o **nome do comprador** na extração.")
    lista = st.session_state.get("lista_pesquisadores", [])
    nome_novo = st.text_input("Nome do pesquisador", key="novo_pesquisador", placeholder="Digite o nome e clique em Adicionar")
    if st.button("➕ Adicionar pesquisador"):
        if nome_novo and nome_novo.strip():
            if nome_novo.strip() not in lista:
                lista.append(nome_novo.strip())
                st.session_state.lista_pesquisadores = lista
                _save_state()
                st.rerun()
        else:
            st.warning("Digite um nome.")
    if lista:
        st.markdown("**Lista atual:**")
        for i, nome in enumerate(lista):
            col1, col2 = st.columns([3, 1])
            with col1:
                st.caption(f"• {nome}")
            with col2:
                if st.button("🗑️ Remover", key=f"remover_{i}"):
                    lista.pop(i)
                    st.session_state.lista_pesquisadores = lista
                    _save_state()
                    st.rerun()

    st.markdown("---")
    st.markdown("**📋 Registro de dúvidas e erros** (últimas entradas)")
    log = []
    lf = _log_duvidas_file()
    if lf.exists():
        try:
            with open(lf, "r", encoding="utf-8") as f:
                log = json.load(f)
        except Exception:
            pass
    if log:
        for entry in reversed(log[-20:]):
            tipo = entry.get("tipo", "")
            emoji = "⚠️" if tipo == "revisar" else "❓" if tipo == "verificar" else "❌"
            st.caption(f"{emoji} [{entry.get('data_hora', '')[:16]}] **{entry.get('arquivo', '?')}** — {entry.get('mensagem', '')}")
    else:
        st.caption("Nenhum registro ainda. As notas com score *revisar* ou *verificar* e erros de processamento são registrados aqui ao longo do tempo.")


def page_sobre():
    st.subheader("ℹ️ Sobre o sistema")
    fluxo_path = Path(__file__).resolve().parent / "pipeline_fluxograma.png"
    if not fluxo_path.exists():
        fluxo_path = Path(__file__).resolve().parent.parent / "assets" / "notas_fiscais_pipeline_flowchart.png"
    col_tex, col_img = st.columns([1, 1])
    with col_tex:
        st.markdown(
            "**📄 Notas Fiscais** extrai dados estruturados de **notas fiscais** usando **OCR** e **modelos de linguagem (IA)**.\n\n"
            "### O que é OCR?\n"
            "**OCR** (*Reconhecimento Óptico de Caracteres*) é a tecnologia que **lê** texto em imagens e PDFs. "
            "Em vez de digitar manualmente, o sistema *enxerga* o documento e transforma o que está escrito em texto editável. "
            "Aqui usamos o **DocTR**, que preserva linhas e espaços para a IA interpretar melhor. "
            "O resultado é um texto <u>organizado</u> por linhas, ideal para a extração dos campos."
        , unsafe_allow_html=True)
    with col_img:
        if fluxo_path.exists():
            st.image(str(fluxo_path), use_container_width=True)
            st.caption("Fluxo do processamento")
    st.markdown("""
    ### 📋 Pipeline

    1. **Upload** — Envio do PDF ou imagem.
    2. **OCR (DocTR)** — Extração do texto com preservação de linhas e espaços.
    3. **Extração (LLM)** — Um modelo de linguagem identifica número da NF, data, empresa, CNPJ, valor, moeda, rubrica, etc.
    4. **Estruturação** — Dados organizados em tabela; filtros por rubrica, data e pesquisador; exportação em CSV.

    ### 🔧 Modelos utilizados

    - **DocTR** — OCR para PDF e imagens.
    - **LLM** — Groq (Llama) ou Hugging Face (Gemma, Qwen, Mistral), configurado no arquivo **.env**. A IA classifica em **rubricas** e preenche os campos.
    """)


def main():
    try:
        st.set_page_config(page_title="Registo Notas Fiscais", page_icon="📄", layout="wide", initial_sidebar_state="expanded")
    except Exception:
        pass  # set_page_config só pode rodar uma vez
    _load_secrets_to_env()
    _load_gcp_secret_manager_to_env()
    inject_css()
    _ensure_dados_dir()
    try:
        init_session_state()
        if not require_login():
            return
        render_sidebar()
        render_header()
        page = st.session_state.get("page", "Início")
        if page == "Início":
            page_inicio()
        elif page == "Processados":
            page_processados()
        elif page == "Revisar":
            page_revisar()
        elif page == "Configurações":
            page_configuracoes()
        else:
            page_sobre()
    except Exception as e:
        st.error(f"**Erro ao carregar o app:** {e}")
        import traceback
        with st.expander("Detalhes técnicos (traceback)"):
            st.code(traceback.format_exc())


if __name__ == "__main__":
    main()
